#!/usr/bin/env python

import argparse
import csv
import math
import os
import statistics
import sys

from collections import Counter

import krippendorff
import numpy as np
import openpyxl
import scipy
import xlsxwriter


HEADER = (
    'labeller_id', 'qa_id', 'story_name', 'section_id', 'section', 'source', 'split',
    'question', 'answer', 'readability', 'relevancy_q', 'relevancy_a'
)

RELABELED_HEADER = ('id', 'section', 'question', 'answer', 'readability_new', 'relevancy_q_new', 'relevancy_a_new')


class StoriesReader:
    def __init__(self):
        self.stories = []
        self.header  = HEADER
        with open(os.path.join('data', 'ACL_StoryQG_Human_Evaluation - Integrated_Results.csv'), newline='') as csvfile:
            reader = csv.DictReader(csvfile, fieldnames=self.header)
            self.original_header = next(reader)
            for row in reader:
                self.stories.append(row)

    def unlabelled_stories_for_labeller(self, labeller):
        unlabelled_stories = []
        for i, story in enumerate(self.stories):
            if int(story['labeller_id']) != int(labeller):
                continue
            new_row = { 'id': i, 'section': story['section'], 'question': story['question'], 'answer': story['answer'] }
            unlabelled_stories.append(new_row)
        return unlabelled_stories

    def relabelled_stories(self):
        rows = []
        with open(os.path.join('data', 'stories_relabelled.csv'), newline='') as csvfile:
            reader = csv.DictReader(csvfile, fieldnames=RELABELED_HEADER)
            next(reader)
            for row in reader:
                rows.append(row)
        return rows


class StoriesWriter:
    def __init__(self, header):
        self.header = header
    
    def write_xlsx_for_labeller(self, labeller, stories):
        workbook = xlsxwriter.Workbook(os.path.join('data', f'new_stories_{labeller}.xlsx'))
        worksheet = workbook.add_worksheet()
        cell_format = workbook.add_format({'text_wrap': True, 'font_name': 'Arial', 'font_size': 10})
        cell_format.set_text_wrap()
        worksheet.set_column_pixels('B:B', 750, cell_format)
        worksheet.set_column_pixels('C:C', 300, cell_format)
        worksheet.set_column_pixels('D:D', 200, cell_format)
        worksheet.set_column_pixels('E:E', 150, cell_format)
        worksheet.set_column_pixels('F:F', 150, cell_format)
        worksheet.set_column_pixels('G:G', 150, cell_format)
        # header row
        worksheet.set_row_pixels(0, 60)
        worksheet.write_row(
            'A1',
            data=['id', 'section', 'question', 'answer', self.header['readability'],self.header['relevancy_q'], self.header['relevancy_a']],
            cell_format=cell_format
        )
        worksheet.freeze_panes(1, 0)
        #
        for i, story in enumerate(stories):
            # 120 characters for a row with 750 px width
            # 1.5 (or 15 if we do not simplify): for correction
            worksheet.set_row_pixels(i + 1, math.ceil(len(story['section'])/12) * 1.5)
            worksheet.write_row(f"A{i + 2}", data=list(story.values()), cell_format=cell_format)
        workbook.close()


def write_files_for_labellers(_args):
    stories_reader = StoriesReader()
    stories_writer = StoriesWriter(stories_reader.original_header)
    for labeller in range(0, 5):
        unlabelled_stories = stories_reader.unlabelled_stories_for_labeller(labeller)
        stories_writer.write_xlsx_for_labeller(labeller, unlabelled_stories)


def anonymize_files(args):
    stories_reader = StoriesReader()
    stories_writer = StoriesWriter(stories_reader.original_header)
    for labeller in range(0, 5):
        stories = []
        workbook = openpyxl.load_workbook(os.path.join('data', f'received_stories_{labeller}.xlsx'))
        worksheet = workbook.active
        for i, row in enumerate(worksheet):
            if not row[1].value:
                break
            if i > 0: # skip header
                stories.append({
                    'id': row[0].value, 'section': row[1].value, 'question': row[2].value, 'answer': row[3].value,
                    'readability': row[4].value, 'relevancy_q': row[5].value, 'relevancy_a': row[6].value
                })
        stories_writer.write_xlsx_for_labeller(labeller, stories)


def merge_labellers_files(_args):
    rows = []
    for labeller in range(0, 5):
        workbook = openpyxl.load_workbook(os.path.join('data', f'new_stories_{labeller}.xlsx'))
        worksheet = workbook.active
        for i, row in enumerate(worksheet):
            if not row[1].value:
                break
            if i > 0: # skip header
                rows.append(row)
    print(len(rows))
    rows = sorted(rows, key=lambda key: int(key[0].value))
    with open(os.path.join('data', f'stories_relabelled.csv'), 'w', newline='') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=RELABELED_HEADER)
        writer.writeheader()
        for row in rows:
            writer.writerow({
                'id': row[0].value, 'section': row[1].value, 'question': row[2].value, 'answer': row[3].value,
                'readability_new': row[4].value, 'relevancy_q_new': row[5].value, 'relevancy_a_new': row[6].value
            })


def combine_labelled_files(_args):
    stories_reader = StoriesReader()
    original_stories = stories_reader.stories
    new_stories = stories_reader.relabelled_stories()
    with open(os.path.join('data', f'stories_combined.csv'), 'w', newline='') as csvfile:
        header = HEADER + ('readability_new', 'relevancy_q_new', 'relevancy_a_new', 'diff_readability', 'diff_relevancy_q', 'diff_relevancy_a')
        writer = csv.DictWriter(csvfile, fieldnames=header)
        writer.writeheader()
        for i, row in enumerate(original_stories):
            new_stories[i].pop('id')
            row.update(new_stories[i])
            # convert to float first because we obtain '5.0' float numbers as strings for some unknown reason
            row['diff_readability'] = abs(int(row['readability']) - int(float(row['readability_new']))) if row['readability'] and row['readability_new'] else ''
            row['diff_relevancy_q'] = abs(int(row['relevancy_q']) - int(float(row['relevancy_q_new']))) if row['relevancy_q'] and row['relevancy_q_new'] else ''
            row['diff_relevancy_a'] = abs(int(row['relevancy_a']) - int(float(row['relevancy_a_new']))) if row['relevancy_a'] and row['relevancy_a_new'] else ''
            writer.writerow(row)


def extract_divergent_examples_base(filter_name, filter_func):
    new_rows = []
    with open(os.path.join('data', f'stories_combined.csv'), newline='') as csvfile:
        reader = csv.DictReader(csvfile)
        for i, row in enumerate(reader):
            if filter_func(row['diff_readability']) or filter_func(row['diff_relevancy_q']) or filter_func(row['diff_relevancy_a']):
                new_rows.append(row)
    with open(os.path.join('data', f'stories_filtered_{filter_name}.csv'), 'w', newline='') as csvfile:
        header = new_rows[0].keys()
        writer = csv.DictWriter(csvfile, fieldnames=header)
        writer.writeheader()
        for row in new_rows:
            writer.writerow(row)


def extract_divergent_examples_geq(threshold):
    extract_divergent_examples_base(f'geq_{threshold}', lambda value: int(value) >= threshold if value else False)

def extract_divergent_examples_eq(threshold):
    extract_divergent_examples_base(f'eq_{threshold}', lambda value: int(value) == threshold if value else False)


def extract_divergent_examples(args):
    if args.op_type == 'eq':
        extract_divergent_examples_eq(args.threshold)
    elif args.op_type == 'geq':
        extract_divergent_examples_geq(args.threshold)


def stories_stats_wrapper(args):
    if not args.system:
        print(
          '[WARNING] No system set - showing global stats for all systems (might not be meaningful).'
          ' Use the option --system {Ours,PAQ,groundtruth}'
        )
    stories_stats(labeller=args.labeller, system=args.system, skip_labellers=args.skip_labellers)


def print_stats(message, data):
    print(f"Mean {message}: {round(statistics.mean(data), 2)}, stdev: {round(statistics.pstdev(data), 2)}")


def stories_stats(labeller=None, system=None, do_print=True, skip_labellers=None):
    if system is not None and do_print:
        print(system)
    with open(os.path.join('data', f'stories_combined.csv'), newline='') as csvfile:
        reader = csv.DictReader(csvfile)
        rows = []
        for i, row in enumerate(reader):
            if labeller is not None:
                if int(row['labeller_id']) != int(labeller):
                    continue
            if skip_labellers is not None and int(row['labeller_id']) in skip_labellers:
                continue
            if system is not None and system != row['source']:
                    continue
            rows.append(row)
        diff_readability = [row['diff_readability'] for row in rows if row['diff_readability']]
        diff_relevancy_q = [row['diff_relevancy_q'] for row in rows if row['diff_relevancy_q']]
        diff_relevancy_a = [row['diff_relevancy_a'] for row in rows if row['diff_relevancy_a']]
        if do_print:
            print("Readability: ", Counter(diff_readability))
            print("Relevancy q: ", Counter(diff_relevancy_q))
            print("Relevancy a: ", Counter(diff_relevancy_a))

        readability = [int(row['readability']) for row in rows if row['readability']]
        relevancy_q = [int(row['relevancy_q']) for row in rows if row['relevancy_q']]
        relevancy_a = [int(row['relevancy_a']) for row in rows if row['relevancy_a']]
        if do_print:
            print_stats("readability", readability)
            print_stats("relevancy q", relevancy_q)
            print_stats("relevancy a", relevancy_a)
            print("Readability: ", Counter(readability))
            print("Relevancy q: ", Counter(relevancy_q))
            print("Relevancy a: ", Counter(relevancy_a))

        readability_new = [int(float(row['readability_new'])) for row in rows if row['readability_new']]
        relevancy_q_new = [int(float(row['relevancy_q_new'])) for row in rows if row['relevancy_q_new']]
        relevancy_a_new = [int(float(row['relevancy_a_new'])) for row in rows if row['relevancy_a_new']]
        if do_print:
            print_stats("readability new", readability_new)
            print_stats("relevancy q new", relevancy_q_new)
            print_stats("relevancy a new", relevancy_a_new)
            print("Readability new: ", Counter(readability_new))
            print("Relevancy q new: ", Counter(relevancy_q_new))
            print("Relevancy a new: ", Counter(relevancy_a_new))
    if system is not None and do_print:
        print("\n")
    if not do_print:
        return readability, relevancy_q, relevancy_a, readability_new, relevancy_q_new, relevancy_a_new


def stats_significance_wrapper(args):
    stats_significance(labeller=args.labeller, skip_labellers=args.skip_labellers)


def stats_significance(labeller=None, skip_labellers=None):
    if labeller is not None:
        print(f"Results for labeller {labeller}")
    ours_readability, ours_relevancy_q, ours_relevancy_a, ours_readability_new, ours_relevancy_q_new, ours_relevancy_a_new = stories_stats(labeller=labeller, skip_labellers=skip_labellers, system='Ours', do_print=False)
    paq_readability, paq_relevancy_q, paq_relevancy_a, paq_readability_new, paq_relevancy_q_new, paq_relevancy_a_new = stories_stats(labeller=labeller, skip_labellers=skip_labellers, system='PAQ', do_print=False)
    gt_readability, gt_relevancy_q, gt_relevancy_a, gt_readability_new, gt_relevancy_q_new, gt_relevancy_a_new = stories_stats(labeller=labeller, skip_labellers=skip_labellers, system='groundtruth', do_print=False)
    print("Readability")
    print("Ours vs PAQ",          scipy.stats.ttest_ind(ours_readability, paq_readability))
    print("Ground truth vs PAQ",  scipy.stats.ttest_ind(gt_readability,   paq_readability))
    print("Ours vs ground truth", scipy.stats.ttest_ind(ours_readability, gt_readability))
    print("Ours vs PAQ (new)",          scipy.stats.ttest_ind(ours_readability_new, paq_readability_new))
    print("Ground truth vs PAQ (new)",  scipy.stats.ttest_ind(gt_readability_new,   paq_readability_new))
    print("Ours vs ground truth (new)", scipy.stats.ttest_ind(ours_readability_new, gt_readability_new))
    print("Ours vs Ours (new)", scipy.stats.ttest_ind(ours_readability, ours_readability_new))
    print("PAQ vs PAQ (new)",   scipy.stats.ttest_ind(paq_readability,  paq_readability_new))
    print("Ground truth vs ground truth (new)", scipy.stats.ttest_ind(gt_readability, gt_readability_new))
    print("---")
    print("Relevancy q")
    print("Ours vs PAQ",          scipy.stats.ttest_ind(ours_relevancy_q, paq_relevancy_q))
    print("Ground truth vs PAQ",  scipy.stats.ttest_ind(gt_relevancy_q,   paq_relevancy_q))
    print("Ours vs ground truth", scipy.stats.ttest_ind(ours_relevancy_q, gt_relevancy_q))
    print("Ours vs PAQ (new)",          scipy.stats.ttest_ind(ours_relevancy_q_new, paq_relevancy_q_new))
    print("Ground truth vs PAQ (new)",  scipy.stats.ttest_ind(gt_relevancy_q_new,   paq_relevancy_q_new))
    print("Ours vs ground truth (new)", scipy.stats.ttest_ind(ours_relevancy_q_new, gt_relevancy_q_new))
    print("Ours vs Ours (new)", scipy.stats.ttest_ind(ours_relevancy_q, ours_relevancy_q_new))
    print("PAQ vs PAQ (new)",   scipy.stats.ttest_ind(paq_relevancy_q,  paq_relevancy_q_new))
    print("Ground truth vs ground truth (new)", scipy.stats.ttest_ind(gt_relevancy_q, gt_relevancy_q_new))
    print("---")
    print("Relevancy a")
    print("Ours vs PAQ",          scipy.stats.ttest_ind(ours_relevancy_a, paq_relevancy_a))
    print("Ground truth vs PAQ",  scipy.stats.ttest_ind(gt_relevancy_a,   paq_relevancy_a))
    print("Ours vs ground truth", scipy.stats.ttest_ind(ours_relevancy_a, gt_relevancy_a))
    print("Ours vs PAQ (new)",          scipy.stats.ttest_ind(ours_relevancy_a_new, paq_relevancy_a_new))
    print("Ground truth vs PAQ (new)",  scipy.stats.ttest_ind(gt_relevancy_a_new,   paq_relevancy_a_new))
    print("Ours vs ground truth (new)", scipy.stats.ttest_ind(ours_relevancy_a_new, gt_relevancy_a_new))
    print("Ours vs Ours (new)", scipy.stats.ttest_ind(ours_relevancy_a, ours_relevancy_a_new))
    print("PAQ vs PAQ (new)",   scipy.stats.ttest_ind(paq_relevancy_a,  paq_relevancy_a_new))
    print("Ground truth vs ground truth (new)", scipy.stats.ttest_ind(gt_relevancy_a, gt_relevancy_a_new))
    

def parse_score(score):
    if score == '':
        score = np.nan
    else:
        score = int(float(score))
    return score


def annotator_agreement(labeller1=None, labeller2=None, label_source1='original', label_source2='original', system=None):
    if labeller1 is not None and labeller1 == labeller2 and label_source1 == label_source2:
        raise ValueError(f'Labellers must be different or with different sources, got {labeller1} and {labeller2}')
    if labeller1 is not None and labeller2 is None:
        print("Warning! labeller2 is None, assuming the same labeller with different source")
        labeller2 = labeller1
        if label_source1 == 'original':
            label_source2 = 'new'
        else:
            label_source2 = 'original'

    if label_source1 == 'original':
        score_suffix1 = ''
    else:
        score_suffix1 = '_new'
    if label_source2 == 'original':
        score_suffix2 = ''
    else:
        score_suffix2 = '_new'

    rows = []
    read_results = {}
    with open(os.path.join('data', f'stories_combined.csv'), newline='') as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            if system is not None and system != row['source']:
                continue
            if labeller1 is not None and int(row['labeller_id']) != labeller1:
                continue
            rows.append(row)
            read_score = row[f'readability{score_suffix1}']
            q_score    = row[f'relevancy_q{score_suffix1}']
            a_score    = row[f'relevancy_a{score_suffix1}']
            #
            read_score = parse_score(read_score)
            q_score    = parse_score(q_score)
            a_score    = parse_score(a_score)
            
            if labeller1 is not None or read_results.get(row['qa_id']) is None:
                read_results[row['qa_id']] = {}
                read_results[row['qa_id']]['labellers']   = [row['labeller_id']]
                read_results[row['qa_id']]['readability'] = [read_score]
                read_results[row['qa_id']]['relevancy_q'] = [q_score]
                read_results[row['qa_id']]['relevancy_a'] = [a_score]
            else:
                if labeller1 is None and read_results.get(row['qa_id']):
                    read_results[row['qa_id']]['labellers'].append(row['labeller_id'])
                    read_results[row['qa_id']]['readability'].append(read_score)
                    read_results[row['qa_id']]['relevancy_q'].append(q_score)
                    read_results[row['qa_id']]['relevancy_a'].append(a_score)

        if labeller1 is not None:
            csvfile.seek(0)
            reader = csv.DictReader(csvfile)
            for row in reader:
                if system is not None and system != row['source']:
                    continue
                read_score = row[f'readability{score_suffix2}']
                q_score    = row[f'relevancy_q{score_suffix2}']
                a_score    = row[f'relevancy_a{score_suffix2}']
                #
                read_score = parse_score(read_score)
                q_score    = parse_score(q_score)
                a_score    = parse_score(a_score)
                
                if read_results.get(row['qa_id']) and int(row['labeller_id']) == labeller2:
                    read_results[row['qa_id']]['labellers'].append(row['labeller_id'])
                    read_results[row['qa_id']]['readability'].append(read_score)
                    read_results[row['qa_id']]['relevancy_q'].append(q_score)
                    read_results[row['qa_id']]['relevancy_a'].append(a_score)

    #print(read_results)
    read_matrix = []
    q_matrix = []
    a_matrix = []
    for k in read_results:
        if len(read_results[k]['labellers']) == 2:
            read_matrix.append(read_results[k]['readability'])
            q_matrix.append(read_results[k]['relevancy_q'])
            a_matrix.append(read_results[k]['relevancy_a'])
        #print(read_results[k])
    read_matrix = np.array(read_matrix).T
    q_matrix = np.array(q_matrix).T
    a_matrix = np.array(a_matrix).T
    print('Read matrix: ')
    print(read_matrix)
    print('Q matrix: ')
    print(q_matrix)
    print('A matrix: ')
    print(a_matrix)
    
    # ordinal or interval or ratio would make the most sense
    level = 'ordinal'
    domain = [1, 2, 3, 4, 5]
    print("Readability:\t\t", krippendorff.alpha(read_matrix, level_of_measurement=level, value_domain=domain))
    print("Question relevancy:\t", krippendorff.alpha(q_matrix, level_of_measurement=level, value_domain=domain))
    print("Answer relevancy:\t", krippendorff.alpha(a_matrix, level_of_measurement=level, value_domain=domain))


def all_aggreements(args):
    if args.labeller1 is not None:
        try:
            annotator_agreement(
                labeller1=args.labeller1,
                labeller2=args.labeller2,
                system=args.system,
                label_source1=args.label_source1,
                label_source2=args.label_source2,
            )
        except ValueError:
            print("[WARNING] No common examples or labeller1 == labeller2")
            pass
        return
    for labeller1 in range(0, 5):
        start = 0 if args.label_source1 != args.label_source2 else labeller1 + 1
        for labeller2 in range(start, 5):
            if labeller1 == labeller2 and args.label_source1 == args.label_source2:
                continue
            try:
                print(f"{labeller1} <-> {labeller2}")
                annotator_agreement(
                    labeller1,
                    labeller2,
                    system=args.system,
                    label_source1=args.label_source1,
                    label_source2=args.label_source2,
                )
            except ValueError:
                pass


def main():
    parser = argparse.ArgumentParser(description="Prepare, process and analyze files related to ReproNLP 2024, Fairytale QA paper.")

    subparsers = parser.add_subparsers(help='The name of the task to run.')


    subparsers.add_parser(
        'write_files_for_labellers',
        help='Write xlsx files for each labeller in data folder: new_stories_{labeller_number}.xlsx',
    ).set_defaults(func=write_files_for_labellers)


    subparsers.add_parser(
        'anonymize_labellers_files',
        help='Rewrite xlsx files provided by each labeller in data folder as received_stories_{labeller_number}.xlsx to new_stories_{labeller_number}.xlsx in order to remove metadata',
    ).set_defaults(func=anonymize_files)


    subparsers.add_parser(
        'merge_labellers_files',
        help=(
                'Place the new_stories_{labeller_number}.xlsx files in the data folder.'
                ' This produces a new file called stories_relabelled.csv'
             ),
    ).set_defaults(func=merge_labellers_files)


    subparsers.add_parser(
        'combine_labelled_files',
        help=(
               'Create the file stories_combined.csv containing the labels from the original experiment,'
               ' the labels from the reproduction study'
               ' and columns with the absolute differences of scores between the two experiments'
             )
    ).set_defaults(func=combine_labelled_files)


    divergent_examples_parser = subparsers.add_parser(
        'extract_divergent_examples',
        help='Extract examples with absolute score differences equal or higher than a threshold'
    )
    divergent_examples_parser.add_argument(
        'threshold', type=int, choices=[2, 3, 4], help='threshold (absolute score difference)'
    )
    divergent_examples_parser.add_argument(
        'op_type', default='eq', nargs='?', choices=['eq', 'geq'], help='comparison_type (default eq)'
    )
    divergent_examples_parser.set_defaults(func=extract_divergent_examples)


    stories_stats_parser = subparsers.add_parser(
        'stories_stats',
        help='Show mean and standard deviation for each system'
    )
    stories_stats_parser.add_argument('--system', choices=['Ours', 'PAQ', 'groundtruth'])
    stories_stats_parser.add_argument('--labeller', choices=[0, 1, 2, 3, 4], type=int, help='Filter results by one labeller')
    stories_stats_parser.add_argument(
        '--skip_labellers', type=int, nargs='+', choices=[0, 1, 2, 3, 4], metavar='N',
        help='Skip results of one or more labellers (choose from 0, 1, 2, 3, 4)'
    )
    stories_stats_parser.set_defaults(func=stories_stats_wrapper)


    stats_significance_parser = subparsers.add_parser(
        'stats_significance',
        help='Show stats significance results presented in the original paper, optionally filtering by labellers'
    )
    stats_significance_parser.add_argument('--labeller', choices=[0, 1, 2, 3, 4], type=int, help='Filter results by one labeller')
    stats_significance_parser.add_argument(
        '--skip_labellers', type=int, nargs='+', choices=[0, 1, 2, 3, 4], metavar='N',
        help='Skip results of one or more labellers (choose from 0, 1, 2, 3, 4)'
    )
    stats_significance_parser.set_defaults(func=stats_significance_wrapper)


    annotator_agreement_parser = subparsers.add_parser(
        'annotator_agreement',
        help='Compute the annotator agreement between all pairs of labellers'
    )
    annotator_agreement_parser.add_argument(
        '--label_source1', default='original', choices=['original', 'new'],
        help='What results to use for the first labeller (default original)'
    )
    annotator_agreement_parser.add_argument(
        '--label_source2', default='original', choices=['original', 'new'],
        help='What results to use for the second labeller (original)'
    )
    annotator_agreement_parser.add_argument(
        '--labeller1', choices=[0, 1, 2, 3, 4], type=int,
        help='First labeller for agreement'
    )
    annotator_agreement_parser.add_argument(
        '--labeller2', choices=[0, 1, 2, 3, 4], type=int,
        help='Second labeller for agreement (default is labeller1 if not set, changes label source to be different)'
    )
    annotator_agreement_parser.add_argument('--system', choices=['Ours', 'PAQ', 'groundtruth'])
    annotator_agreement_parser.set_defaults(func=all_aggreements)
    

    args = parser.parse_args()
    if len(sys.argv) > 1:
        args.func(args)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()


exit()

# 1)
# python main.py write_files_for_labellers

# 2)
# # move/add newly labelled files to data/ directory

# 3)
# python main.py anonymize_labellers_files

# 4)
# python main.py merge_labellers_files
# python main.py combine_labelled_files

# 5)
# python main.py extract_divergent_examples 4
# python main.py extract_divergent_examples 3

# 6)
# python main.py stories_stats --system Ours
# python main.py stories_stats --system PAQ
# python main.py stories_stats --system groundtruth

# python main.py stories_stats --system Ours --skip_labellers 1
# python main.py stories_stats --system PAQ --skip_labellers 1
# python main.py stories_stats --system groundtruth --skip_labellers 1

# 7)
# python main.py stats_significance
# python main.py stats_significance --skip_labellers 1

# 8)
# python main.py annotator_agreement
# python main.py annotator_agreement --label_source1=new --label_source2=new
# python main.py annotator_agreement --label_source1=new
# python main.py annotator_agreement --system=Ours --label_source1=new --label_source2=new
# python main.py annotator_agreement --labeller1=0 --labeller1=1


